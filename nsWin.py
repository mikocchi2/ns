import io
from openpyxl import *
from datetime import *
import os
import shutil
import configparser

from gmailapilib import *

# star imports su BAD
# zbog from openpyxl import * mi ne radi open()



max_results = 10

radnici = []
radnici_za_racun = []

d1 = None
d2 = None
kol = 0

broj_rac = None

ns_mail = "prodaja@omnisplus.rs"
ac = "aleksandarvasiljevic11@gmail.com"

message_id = None


SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/gmail.send"
]




#  invoice maker
def no_gui():
    
    load_update_configs()
    delta_no_gui()
    upisi_datum()
    broj_racuna()
    priprema()
    cuvaj()
    pdf_path = pdf()

    return pdf_path
#
def delta_no_gui():
    global d1, d2
    # Convert string dates to datetime objects for calculation
    format = "%d.%m.%Y"
    dat1 = datetime.strptime(d1, format).date()
    dat2 = datetime.strptime(d2, format).date()
    delta = dat2 - dat1
    print(f"Količina dana: {delta.days}")
    global kol
    kol = delta.days
#
def load_update_configs():
    def update_price():
        # write to template
        wb = load_workbook("ns_template.xlsx")
        ws = wb.active
        print(ukupna_cena)
        ws['D22'] = ukupna_cena
        wb.save("ns_template.xlsx")
        wb.close
    
    # OPEN FILE
    config = configparser.ConfigParser()
    config.read('config.ini')

    # load vars
    cena = int(config['DEFAULT']['cena']) 
    boravisna = int(config['DEFAULT']['boravisna'])
    ukupna_cena = cena+boravisna
    global broj_rac
    broj_rac = int(config["DEFAULT"]['broj_racuna'])

    update_price()
# 
def upisi_datum():
    wb = load_workbook("ns_template.xlsx")
    ws = wb.active
    # global vars strasno
    d1_sliced = str(d1)[:6]
    datum = d1_sliced + " - " + d2 + '.'
    ws['A16'] = datum
    wb.save('ns_template.xlsx')
    wb.close() 
    print(datum)
# 
def broj_racuna():

    def update_br_rac():
        config = configparser.ConfigParser()
        config.read('config.ini')
        # Since we're working with the DEFAULT section, set the value directly
        config['DEFAULT']['broj_racuna'] = str(broj_rac + 1)
        with io.open('config.ini', 'w') as configfile:
            config.write(configfile)

    wb_ns = load_workbook("ns_template.xlsx")
    ws_ns = wb_ns.active

    # samo 23,24...
    curr_godina = str(datetime.now().year)[2:]
    # formatira ga lepo
    global broj_racuna_string
    if broj_rac < 100: broj_racuna_string = "br.0" + str(broj_rac) + "/" + curr_godina
    else:              broj_racuna_string = "br." + str(broj_rac) + "/" + curr_godina

    ws_ns["A14"].value = broj_racuna_string

    
    wb_ns.save("ns_template.xlsx")
    #wb_b_rac.close()
    wb_ns.close()
    update_br_rac()
#                                     pravi tmp ns_racun.xlsx
def priprema():

    wb = load_workbook("ns_template.xlsx")
    ws = wb.active
    ws_rows = ws.iter_rows(min_row=22, max_row=32, min_col=1, max_col=5)
    
    counter_i = 0 # za sve radnike u radnici_za_racun
    for row in ws_rows:
        if counter_i <= len(radnici_za_racun)-1:
            row[0].value = radnici_za_racun[counter_i]
            row[2].value = str(kol)
            counter_i += 1
        else:
            break
    
    # 22 - start row, 7 - max input rows, minus len popunjenih
    index = 22 + len(radnici_za_racun)
    brisanje = 20 - len(radnici_za_racun)
    ws.delete_rows(index, brisanje)  
    wb.save('ns_racun.xlsx')
    print("priprema over")
    wb.close()
# 
def cuvaj():

    # Get the current working directory
    currentdir = os.getcwd()
    # Path to the source file
    currentfile = os.path.join(currentdir, "ns_racun.xlsx")
    
    # Assuming 'broj_racuna_string' is defined globally and contains a string like '123abc'
    global ime_fajla
    ime_fajla = broj_racuna_string[3:]
    
    # TMP!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    ime_fajla = ime_fajla.replace("/", "-")


    # Create the destination path, ensuring to use os.path.join for compatibility
    racuni_dir = os.path.join(currentdir, "racuni")
    # Check if the 'racuni' directory exists, create if not
    if not os.path.exists(racuni_dir):
        os.makedirs(racuni_dir)
    
    # Construct the full path to the destination file
    to = os.path.join(racuni_dir, ime_fajla + ".xlsx")
    
    # Copy the file
    shutil.copy(currentfile, to)
 
    print("uradio")
# 
def pdf():
        
        import win32com.client

        o = win32com.client.Dispatch("Excel.Application")
        o.Visible = False
        o.DisplayAlerts = False
        
        xlsx_path = os.getcwd() + "/racuni/" + ime_fajla + ".xlsx"
        pdf_create_path = os.getcwd() + "/racuni/pdf"
        pdf = pdf_create_path + '/' + ime_fajla + ".pdf"

        print(xlsx_path)
        wb = o.Workbooks.Open(xlsx_path) 
        wb.ActiveSheet.ExportAsFixedFormat(0, pdf)
        o.Quit()

        return pdf






# mail-reader
# ovde je a ne u gmailapilib jer imam global variables, VERY GOODO PURACUTICU
def parse_email_body(service, labelName,from_email,user_id='me'):

    parsed_thread_ids = []
    

    # query
    label_Id = get_labelId(service,labelName)
    query = f"from:{from_email}" if from_email else None

    # [ emails ]
    results = service.users().messages().list(userId=user_id, labelIds=[label_Id], q=query,maxResults=max_results).execute()
    messages = results.get('messages', [])

    counter = 1
    for msg in messages:
        print(len(messages))
        print(msg['id'])
        #is_read = 'UNREAD' not in message.get('labelIds', [])
        
        
        global message_id
        # check for replies and skip them (it fucks with the parser)
        message_id,tid = msg['id'],msg['threadId']
        if tid in parsed_thread_ids: continue
        else: parsed_thread_ids.append(tid)

        # message is legit
        message = get_mime_message(service, user_id, msg['id'])
    
        # msg is either single or multipart parse the body accordingly
        if message.is_multipart():
            for part in message.walk():
                encoding = get_charset(part=part)
                content_type = part.get_content_type()
                content_disposition = str(part.get('Content-Disposition'))
                if content_type == 'text/plain' and 'attachment' not in content_disposition:
                    msgbdy = part.get_payload(decode=True).decode(encoding=encoding)
        else:
            msgbdy = message.get_payload(decode=True).decode(encoding=encoding)

        # process the msgbdy
        response = json.loads(process_mail_gpt(msgbdy))

        global d1,d2,radnici_za_racun
        d1 = response['OD']
        d2 = response['DO']
        radnici_za_racun = response['clients']
        print(d1,d2,radnici_za_racun)
        return d1,d2,radnici_za_racun





# i dont even know man   
if __name__ == "__main__":

    service = get_service()
    sendermail = ns_mail
    parse_email_body(service=service,labelName='ns',from_email=sendermail,user_id='me')
    
    old_label = get_labelId(service,'ns')
    new_label = get_labelId(service,'nsold')

    pdf_file_path = no_gui()

    message = create_pdf_mail_reply(message_id, 'me', "", pdf_file_path)
    send_message(service, "me", message)
    move_message_to_label(service,'me',message_id,old_label,new_label)